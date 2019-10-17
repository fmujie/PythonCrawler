# coding=utf-8
from selenium import webdriver
from lxml import etree
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import re
import time


class LaGouSpider(object):
    driver_path = r"D:\Software_installation_program\chromedriver\chromedriver_win32\chromedriver.exe"

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=LaGouSpider.driver_path)
        self.url = 'https://www.lagou.com/jobs/list_python?labelWords=&fromSearch=true&suginput='
        self.datas = []

    def run(self):
        self.driver.get(self.url)
        source = self.driver.page_source
        self.parse_list_page(source)

    def parse_list_page(self, source):
        html = etree.HTML(source)
        links = html.xpath("//a[@class='position_link']/@href")
        for link in links:
            self.request_detail_page(link)
            time.sleep(3)

    def request_detail_page(self, url):
        self.driver.get(url)
        source = self.driver.page_source
        self.parse_detail_page(source)

    def parse_detail_page(self, source):
        html = etree.HTML(source)
        position_name = html.xpath("//h2[@class='name']/text()")
        # print(position_name)
        job_request_span = html.xpath("//dd[@class='job_request']//span")
        salary = job_request_span[0].xpath(".//text()")[0].strip()
        # print(salary)
        dataDictionary = {
            'name': position_name,
            'salary': salary
        }
        self.datas.append(dataDictionary)
        print(dataDictionary)
        print('=' * 30)

    def process_workbook(self, filename):
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            cell.value = self.datas[row - 2]
        wb.save(filename)

    # def process_workbook(filename):
    #     wb = xl.load_workbook(filename)
    #     sheet = wb['Sheet1']
    #     for row in range(2, sheet.max_row + 1):
    #         cell = sheet.cell(row, 2)
    #         cell.value = "shsh"
    #         print(cell.value)
    #     wb.save(filename)


if __name__ == '__main__':
    spider = LaGouSpider()
    spider.run()
    # print(self.datas)
