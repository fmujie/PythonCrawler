import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# dic_data = {
#     'legal_status_update_time': '',
#
# }
# a = 'safsdf'
# dic_data['legal_status_update_time'] = a
# print(dic_data['legal_status_update_time'])
# dic_data.clear()
# print(dic_data['legal_status_update_time'])
# test = {
#     'name': ['fmu', 'fangmu'],
#     'age': '18',
# }
# list = ['www', 'google', 'com']
# list1 = ['www']
# str4 = ";".join(list)
# print(str4)

# print(test['name'][0])
# a = 'sdff\t\n'

# print(a.strip())

# def list_to_str(listInfo):
#     str = ";".join(listInfo)
#     return str

# def process_workbook(filename):
#     wb = xl.load_workbook(filename)
#     # name_list = wb.sheetnames
#     # print(name_list)
#     # sheet = wb['Sheet1']
#     active_sheet = wb.active
#     # print(active_sheet)
#     for row in range(2, active_sheet.max_row + 1):
#         cell = active_sheet.cell(row, 3)
#         cell.value = list1[0] #test['name'][0]
#         print(cell.value)
#         print("="*30)
#     wb.save(filename)


# process_workbook('test.xlsx')
# print(list1)
# print(list_to_str(list1))


index = 0

def foo():
    global index
    index += 1
    return index

print(index)
foo()
print(index)
foo()
print(index)