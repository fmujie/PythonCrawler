# coding=utf-8
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from lxml import etree
import warnings
import time
import re


class ReptilianRobot(object):

    driver_path = r"D:\Software_installation_program\chromedriver\chromedriver_win32\chromedriver.exe"
    login_url = 'https://analytics.zhihuiya.com/'
    userName = ''
    passWord = ''
    index = 2

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=ReptilianRobot.driver_path)
        self.datas = []

    def login(self):
        self.driver.get(self.login_url)
        self.driver.set_window_size(1500, 1000)
        time.sleep(1.5)
        userNameInput = self.access_to_elements("//input[@id='acount']")
        passWordInput = self.access_to_elements("//input[@id='password']")
        userNameInput.send_keys(self.userName)
        time.sleep(1.5)
        passWordInput.send_keys(self.passWord)
        submitBtn = self.access_to_elements("//button[@id='log-button']")
        submitBtn.click()

    def fill_patent_number(self, serial_number):
        time.sleep(1.5)
        self.driver.implicitly_wait(1)
        searchInputBox = self.access_to_elements("//div[@id='syncWidthElementId1']/input")
        time.sleep(1.5)
        searchInputBox.send_keys(serial_number)
        # time.sleep(1.5)
        # searchInputBox = self.access_to_elements("//div[@id='syncWidthElementId1']/input")
        # actions = ActionChains(self.driver)
        # actions.move_to_element(searchInputBox)
        # time.sleep(1.5)
        # actions.send_keys(serial_number)
        # actions.perform()

    def search_patent_number(self):
        time.sleep(1.5)
        searchBtn = self.access_to_elements("//button[@class='s-search']")
        searchBtn.click()
        # searchBtn = self.access_to_elements("//button[@class='s-search']")
        time.sleep(1.5)
        # actions = ActionChains(self.driver)
        # actions.move_to_element(searchBtn)
        # actions.click(searchBtn)
        # actions.perform()

    def parsing_search_return_list(self):
        time.sleep(1.5)
        element = self.access_to_elements("//div[@class='pn-cell']/a[@class='pn-cell-popover']")
        actions = ActionChains(self.driver)
        actions.move_to_element(element)
        actions.click(element)
        actions.perform()

    def switch_pages_to_details(self):
        self.driver.switch_to.window(self.driver.window_handles[1])
        time.sleep(1.5)
        self.driver.implicitly_wait(1)
        source = self.driver.page_source
        return source

    def parse_details_page(self, source, serial_number):
        print(serial_number)
        html = etree.HTML(source)
        application_number = html.xpath(
            "//div[@class='table-row'][3]/div[@class='origin content']/div[@class='view-item view-30 view-30-spc']/div[@class='table-text cell']/text()")
        application_number_toStr = self.list_to_str(application_number)

        title = html.xpath("//p[@class='pn-title']/text()")
        title_toStr = self.list_to_str(title)

        filing_publication = html.xpath(
            "//div[@class='origin content']/div[@class='view-item view-30 view-30-label']/div[@class='table-text cell']/text()")
        filing_publication_toStr = self.list_to_str(filing_publication)

        IPC_classification_number = html.xpath(
            "//div[@class='table-row'][8]/div[@class='origin content']/div[@class='table-text cell']/div[@class='address-row']/span[@class='row']/a[@class='item tooltipstered']/text()")
        IPC_classification_number_toStr = self.list_to_str(IPC_classification_number)

        current_aplicant = html.xpath(
            "//div[@class='table-row'][4]/div[@class='origin content']/div[@class='table-text cell fixed']/div[@class='address-row']/span[@class='row is-block']/a[@class='name item tooltipstered']/text()")
        current_aplicant_toStr = self.list_to_str(current_aplicant)

        original_applicant = html.xpath(
            "//div[@class='table-row'][6]/div[@class='origin content']/div[@class='table-text cell fixed']/div[@class='address-row']/span[@class='row is-block']/a[@class='name item tooltipstered']/text()")
        original_applicant_toStr = self.list_to_str(original_applicant)

        current_aplicant_address = html.xpath(
            "//div[@class='table-row'][5]/div[@class='origin content']/div[@class='table-text cell fixed']/div[@class='address-row']/span[@class='row is-block']/p[2]/span[@class='address item']/text()")
        current_aplicant_address_toStr = self.list_to_str(current_aplicant_address)

        inventor = html.xpath(
            "//div[@class='table-row'][11]/div[@class='origin content']/div[@class='table-text cell']/div[@class='address-row']/span[@class='row is-block']/span[@class='name item']/text()")
        inventor_toStr = self.list_to_str(inventor)

        agency = html.xpath(
            "//div[@class='table-row'][13]/div[@class='origin content']/div[@class='table-text cell']/div[@class='address-row']/span[@class='row is-block']/span[@class='name item']/text()")
        agency_toStr = self.list_to_str(agency)
        '''
        附图
        '''
        # time.sleep(1.5)
        self.switch_to_drawing()
        time.sleep(1.5)
        source = self.driver.page_source
        html = etree.HTML(source)
        drawing_src = html.xpath("//div[@class='image-wrap text-figure__full']/img[@class='image-img src-img']/@src")
        drawing_src_toStr = self.list_to_str(drawing_src)

        # time.sleep(1.5)
        self.switch_to_graphic()
        time.sleep(1.5)
        source = self.driver.page_source
        html = etree.HTML(source)
        abstract = html.xpath("//div[@class='pt-section-abst pt-section']/div[@class='pt-section__body']/text()")
        abstract_toStr = self.list_to_str(abstract)

        # first_authority_request = html.xpath("//div[@class='pt-section-clms pt-section']/div[@class='pt-section__body']/span/p/text()")[0]

        authority_request = html.xpath("//div[@class='pt-section-clms pt-section']/div[@class='pt-section__body']/span/p/text()")
        # first_authority_request = authority_request[0]
        authority_request_toStr = self.list_to_str(authority_request)

        self.switch_to_legal()
        time.sleep(1.5)
        source = self.driver.page_source
        html = etree.HTML(source)
        legal_status = html.xpath("//div[@class='tag-container']/span[@class='tag legal-2']/text()")
        legal_status_toStr = (self.list_to_str(legal_status)).strip()

        legal_status_update_time = html.xpath("//tbody[1]/tr[1]/th[@class='date']/text()")
        legal_status_update_time_toStr = self.list_to_str(legal_status_update_time)
        dict_data = {
            'application_number': application_number_toStr,
            'title': title_toStr,
            'filing_publication': filing_publication_toStr,
            'IPC_classification_number': IPC_classification_number_toStr,
            'current_aplicant': current_aplicant_toStr,
            'original_applicant': original_applicant_toStr,
            'current_aplicant_address': current_aplicant_address_toStr,
            'inventor': inventor_toStr,
            'agency': agency_toStr,
            'drawing_src': drawing_src_toStr,
            'abstract': abstract_toStr,
            # 'first_authority_request': first_authority_request,
            'authority_request': authority_request_toStr,
            'legal_status': legal_status_toStr,
            'legal_status_update_time': legal_status_update_time_toStr
        }
        print(dict_data)
        # print(dict_data['application_number_toStr'])
        # print(dict_data['title_toStr'])
        # print(dict_data['filing_publication_toStr'])
        # print(dict_data['IPC_classification_number_toStr'])
        # print(dict_data['current_aplicant_toStr'])
        # print(dict_data['original_applicant_toStr'])
        # print(dict_data['current_aplicant_address_toStr'])
        # print(dict_data['inventor_toStr'])
        # print(dict_data['agency_toStr'])
        # print(dict_data['drawing_src_toStr'])
        # print(dict_data['abstract_toStr'])
        # print(dict_data['first_authority_request'])
        # print(dict_data['authority_request_toStr'])
        # print(dict_data['legal_status_toStr'])
        # print(dict_data['legal_status_update_time_toStr'])
        self.storage_data(dict_data, 'work.xlsx', serial_number)
        del dict_data

    def storage_data(self, dic, filename, serial_number):
        global index
        wbs = xl.load_workbook(filename)
        active_sheet = wbs.active
        cell_serial_number = active_sheet.cell(self.index, 2)
        cell_serial_number.value = serial_number
        cell_application_number = active_sheet.cell(self.index, 3)
        cell_application_number.value = dic['application_number']
        cell_title_toStr = active_sheet.cell(self.index, 4)
        cell_title_toStr.value = dic['title']
        cell_filing_publication = active_sheet.cell(self.index, 5)
        cell_filing_publication.value = dic['filing_publication']
        cell_IPC_classification_number = active_sheet.cell(self.index, 6)
        cell_IPC_classification_number.value = dic['IPC_classification_number']
        cell_current_aplicant = active_sheet.cell(self.index, 7)
        cell_current_aplicant.value = dic['current_aplicant']
        cell_original_applicant = active_sheet.cell(self.index, 8)
        cell_original_applicant.value = dic['original_applicant']
        cell_current_aplicant_address = active_sheet.cell(self.index, 9)
        cell_current_aplicant_address.value = dic['current_aplicant_address']
        cell_inventor = active_sheet.cell(self.index, 10)
        cell_inventor.value = dic['inventor']
        cell_agency = active_sheet.cell(self.index, 11)
        cell_agency.value = dic['agency']
        cell_drawing_src = active_sheet.cell(self.index, 12)
        cell_drawing_src.value = dic['drawing_src']
        cell_abstract = active_sheet.cell(self.index, 13)
        cell_abstract.value = dic['abstract']
        # cell_first_authority_request = active_sheet.cell(self.index, 14)
        # cell_first_authority_request.value = dic['first_authority_request']
        cell_authority_request = active_sheet.cell(self.index, 14)
        cell_authority_request.value = dic['authority_request']
        cell_legal_status = active_sheet.cell(self.index, 15)
        cell_legal_status.value = dic['legal_status']
        cell_legal_status_update_time = active_sheet.cell(self.index, 16)
        cell_legal_status_update_time.value = dic['legal_status_update_time']
        wbs.save(filename)
        self.index += 1
        return self.index

    def switch_to_graphic(self):
        time.sleep(1.5)
        click_graphic = self.access_to_elements("//div[@class='tab-mode mode-search']/span[2]/a[@id='dual']")
        click_graphic.click()

    def switch_to_legal(self):
        time.sleep(1.5)
        click_legal = self.access_to_elements("//div[@class='tab-mode mode-search']/span[5]/a[@id='legal']")
        click_legal.click()

    def switch_to_drawing(self):
        time.sleep(1.5)
        element = self.access_to_elements("//div[@class='side-nav']/ul/li[@id='thumb']")
        actions = ActionChains(self.driver)
        actions.move_to_element(element)
        actions.click(element)
        actions.perform()

    def close_back(self):
        self.driver.close()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.back()

    def determine_nor_null(self):
        self.driver.implicitly_wait(1)
        searchInputBox = self.access_to_elements("//div[@id='syncWidthElementId1']/input")
        searchInputBox.clear()

    def access_to_elements(self, XPathValue):
        try:
            element = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, XPathValue))
            )
            return element
        except TimeoutException as ex:
            print(str(ex) + "TimeOutException")

    def list_to_str(self, listInfo):
        str = " ".join(listInfo)
        return str

    def process_workbook(self, filename):
        wb = xl.load_workbook(filename)
        active_sheet = wb.active
        for row in range(2, active_sheet.max_row + 1):
            serial_number = active_sheet.cell(row, 2)
            print(serial_number.value)
            self.repetitive_operation(serial_number.value)
        wb.save(filename)

    def repetitive_operation(self, serial_number):
        self.fill_patent_number(serial_number)
        self.search_patent_number()
        self.parsing_search_return_list()
        self.switch_pages_to_details()
        source = self.switch_pages_to_details()
        self.parse_details_page(source, serial_number)
        self.close_back()
        self.determine_nor_null()


if __name__ == '__main__':
    test = ReptilianRobot()
    test.login()
    test.process_workbook('test.xlsx')
